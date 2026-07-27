"""
build_rating_checklist.py

생성된 퀴즈(results/generated_quizzes.json)를 바탕으로, 채점자 3명이 각자
독립적으로 채점할 수 있는 엑셀 체크리스트를 만든다.

채점 기준 2개 (0/1로 채점):
  - 명확성: 질문이 애매하지 않고 뭘 묻는지 분명한가
  - 교육적가치: 신입사원이 실무에서 실제로 알아야 할 핵심 내용을 다루는가

3명이 서로 다른 파일(rater1/2/3)을 독립적으로 채점하게 해서, 서로의
채점 결과에 영향받지 않는 블라인드 평가가 되도록 함.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CRITERIA = ["명확성", "교육적가치"]


def load_quizzes(path: str | Path) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _choices_text(quiz: dict) -> str:
    choices = quiz.get("choices") or []
    if not choices:
        return "(단답형 - 보기 없음)"
    return "\n".join(f"{i}. {c}" for i, c in enumerate(choices, start=1))


def build_checklist(quizzes: list[dict], output_path: str | Path, rater_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "채점표"

    headers = [
        "quiz_id", "질문", "보기", "정답", "해설", "근거", "슬라이드",
        *[f"{c} (0/1)" for c in CRITERIA],
        "코멘트(선택)",
    ]
    ws.append(headers)

    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, quiz in enumerate(quizzes, start=2):
        answer = quiz.get("answer", "")
        choices = quiz.get("choices") or []
        answer_text = (
            f"{answer}. {choices[answer - 1]}"
            if isinstance(answer, int) and 1 <= answer <= len(choices)
            else str(answer)
        )
        row = [
            quiz.get("quiz_id", ""),
            quiz.get("question", ""),
            _choices_text(quiz),
            answer_text,
            quiz.get("explanation", ""),
            quiz.get("evidence", ""),
            quiz.get("source", {}).get("slide", ""),
            "", "",  # 명확성, 교육적가치 -- 채점자가 직접 채워넣는 빈 칸
            "",
        ]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=i, column=col)
            c.font = normal_font = Font(name="맑은 고딕", size=10)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=(col in (2, 3, 5, 6)))

    widths = {1: 12, 2: 30, 3: 25, 4: 20, 5: 28, 6: 28, 7: 8, 8: 12, 9: 12, 10: 20}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(quizzes) + 1}"

    # 채점 안내 시트
    guide = wb.create_sheet("채점 안내")
    guide["A1"] = f"채점자: {rater_name}"
    guide["A1"].font = Font(bold=True, size=13)
    instructions = [
        "",
        "각 문항에 대해 아래 2개 기준을 0 또는 1로만 채점해주세요.",
        "(다른 채점자와 상의하지 말고 독립적으로 채점 부탁드립니다)",
        "",
        "1) 명확성 (0/1)",
        "   1 = 질문이 무엇을 묻는지 분명하고, 보기 중 정답이 명확히 구분됨",
        "   0 = 질문이 모호하거나, 문장이 어색해서 뭘 묻는지 헷갈림",
        "",
        "2) 교육적가치 (0/1)",
        "   1 = 신입사원이 실무에서 실제로 알아야 할 핵심 내용을 다룸",
        "   0 = 지엽적이거나 실무와 관련 없는 사소한 디테일을 다룸",
        "",
        "코멘트 칸은 선택사항입니다 (특이사항 있을 때만 적어주세요).",
    ]
    for i, line in enumerate(instructions, start=2):
        guide.cell(row=i, column=1, value=line).font = Font(name="맑은 고딕", size=11)
    guide.column_dimensions["A"].width = 70

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    import sys

    quizzes_path = sys.argv[1] if len(sys.argv) > 1 else r"C:\team-05-project\src\generated_quizzes.json"
    quizzes = load_quizzes(quizzes_path)
    print(f"퀴즈 {len(quizzes)}개 로딩됨")

    for rater in ["rater1", "rater2", "rater3"]:
        output = f"results/checklist_{rater}.xlsx"
        build_checklist(quizzes, output, rater_name=rater)
        print(f"생성됨: {output}")
