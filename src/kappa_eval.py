"""
kappa_eval.py

채점자 3명이 채운 엑셀(checklist_rater1/2/3.xlsx)과 AI 채점 결과
(ai_judgments.json)를 읽어서:
  1) 사람 3명끼리의 일치도 -> Fleiss' Kappa
  2) AI vs 사람(다수결)의 일치도 -> Cohen's Kappa
을 "명확성", "교육적가치" 각각 따로 계산한다.

라이브러리:
    pip install openpyxl statsmodels scikit-learn
"""
from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from sklearn.metrics import cohen_kappa_score
from statsmodels.stats.inter_rater import fleiss_kappa

CRITERIA = ["명확성", "교육적가치"]


def load_ratings(xlsx_path: str | Path) -> dict[str, dict[str, int]]:
    """엑셀 체크리스트를 읽어서 {quiz_id: {"명확성": 0/1, "교육적가치": 0/1}} 형태로 반환."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["채점표"]
    header = [cell.value for cell in ws[1]]
    col_idx = {name: i for i, name in enumerate(header)}

    ratings = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        quiz_id = row[col_idx["quiz_id"]]
        if not quiz_id:
            continue
        entry = {}
        for crit in CRITERIA:
            value = row[col_idx[f"{crit} (0/1)"]]
            if value is None or value == "":
                raise ValueError(
                    f"{xlsx_path}의 {quiz_id} 행에서 '{crit}' 값이 비어있습니다. "
                    f"채점이 다 끝났는지 확인하세요."
                )
            entry[crit] = int(value)
        ratings[str(quiz_id)] = entry
    return ratings


def majority_vote(values: list[int]) -> int:
    """짝수 동점 방지를 위해 홀수(3명) 기준. 동점이면 보수적으로 0(부적합) 처리."""
    counts = Counter(values)
    if len(counts) == 1:
        return values[0]
    most_common = counts.most_common()
    if most_common[0][1] > most_common[1][1]:
        return most_common[0][0]
    return 0  # 동점(불가능해야 하나 방어적으로)


def compute_fleiss_kappa(rater_ratings: list[dict[str, dict[str, int]]], criterion: str, quiz_ids: list[str]) -> float:
    """rater_ratings: [rater1_dict, rater2_dict, rater3_dict]
    각 문항(row) x 카테고리(0 또는 1) 개수 행렬을 만들어 Fleiss' Kappa 계산."""
    table = []
    for qid in quiz_ids:
        values = [r[qid][criterion] for r in rater_ratings]
        count_0 = values.count(0)
        count_1 = values.count(1)
        table.append([count_0, count_1])
    return fleiss_kappa(np.array(table), method="fleiss")


def compute_cohen_kappa(ai_values: list[int], human_majority_values: list[int]) -> float:
    if len(set(ai_values)) == 1 and len(set(human_majority_values)) == 1:
        # 둘 다 전부 같은 값만 나온 경우 -- sklearn이 nan을 반환하는 특수 케이스
        return float("nan")
    return cohen_kappa_score(ai_values, human_majority_values)


def raw_agreement(a: list[int], b: list[int]) -> float:
    return sum(1 for x, y in zip(a, b) if x == y) / len(a)


def main(results_dir: str = "results") -> None:
    results_dir = Path(results_dir)

    rater_files = [results_dir / f"checklist_rater{i}.xlsx" for i in (1, 2, 3)]
    rater_ratings = [load_ratings(f) for f in rater_files]

    with open(results_dir / "ai_judgments.json", encoding="utf-8") as f:
        ai_raw = json.load(f)
    ai_ratings = {
        item["quiz_id"]: item for item in ai_raw
        if item.get("명확성") is not None and item.get("교육적가치") is not None
    }

    # 3명 다 채점했고, AI 채점도 성공한 quiz_id만 대상으로 함
    quiz_ids = sorted(
        set(rater_ratings[0]) & set(rater_ratings[1]) & set(rater_ratings[2]) & set(ai_ratings)
    )
    print(f"평가 대상 문항 수: {len(quiz_ids)}개\n")
    if len(quiz_ids) < 10:
        print("[주의] 표본이 너무 적습니다. Kappa 값이 불안정할 수 있어요 (최소 30개 이상 권장).\n")

    print("=" * 70)
    print(f"{'기준':12s} | {'사람-사람 raw일치율':>16s} | {'Fleiss Kappa':>13s} | "
          f"{'AI-사람 raw일치율':>15s} | {'Cohen Kappa':>12s}")
    print("-" * 70)

    for criterion in CRITERIA:
        r1 = [rater_ratings[0][q][criterion] for q in quiz_ids]
        r2 = [rater_ratings[1][q][criterion] for q in quiz_ids]
        r3 = [rater_ratings[2][q][criterion] for q in quiz_ids]
        ai = [ai_ratings[q][criterion] for q in quiz_ids]
        human_maj = [majority_vote([a, b, c]) for a, b, c in zip(r1, r2, r3)]

        pair_agreements = [raw_agreement(r1, r2), raw_agreement(r1, r3), raw_agreement(r2, r3)]
        human_raw = sum(pair_agreements) / len(pair_agreements)
        fleiss = compute_fleiss_kappa([rater_ratings[0], rater_ratings[1], rater_ratings[2]], criterion, quiz_ids)

        ai_raw_agree = raw_agreement(ai, human_maj)
        cohen = compute_cohen_kappa(ai, human_maj)

        print(f"{criterion:12s} | {human_raw:>15.3f}  | {fleiss:>12.3f}  | "
              f"{ai_raw_agree:>14.3f}  | {cohen:>11.3f}")

    print("=" * 70)
    print("""
해석 기준 (Kappa 값):
  0.8 이상    = 거의 완벽하게 일치
  0.6 ~ 0.8   = 상당히 일치
  0.4 ~ 0.6   = 보통 수준 일치
  0.4 미만    = 낮은 일치 (신뢰하기 어려움)

[주의] raw 일치율은 높은데 Kappa가 낮게 나오면, 대부분 같은 값(예: 다 1)으로
쏠려서 생기는 통계적 역설(Kappa paradox)일 수 있습니다. 이 경우 raw 일치율과
함께 해석하고, 채점 결과의 분포(0/1 비율)를 같이 확인해보세요.
""")


if __name__ == "__main__":
    import sys

    main(sys.argv[1] if len(sys.argv) > 1 else "results")
